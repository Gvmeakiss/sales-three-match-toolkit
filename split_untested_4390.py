# -*- coding: utf-8 -*-
"""
将 4390 的 Untested 单文件拆分为多个 xlsx，便于加载。

拆分规则：
- 文件 1（_分片_1.xlsx）：仅订单 + 仅订单及发货单 + 仅发货单 + 负开票冲帐_1
- 文件 2~7（_分片_2 ~ _分片_7.xlsx）：负开票冲帐_2 ~ 负开票冲帐_7

使用 openpyxl read_only/write_only 流式读写，避免大文件内存溢出。

用法: python3 split_untested_4390.py
"""

import os
import sys
import glob

script_dir = os.path.dirname(os.path.abspath(__file__))
if script_dir not in sys.path:
    sys.path.insert(0, script_dir)

from openpyxl import load_workbook
from config import OUTPUT_ROOT, OUTPUT_PREFIX
from utils.path_utils import ensure_dir


def _stream_sheet_to_file(src_wb, sheet_name, out_path, batch_log=500000):
    """流式复制单个 sheet 到新文件，避免全量加载"""
    ws_src = src_wb[sheet_name]
    from openpyxl import Workbook
    wb_dst = Workbook(write_only=True)
    ws_dst = wb_dst.create_sheet(sheet_name[:31])
    n = 0
    for row in ws_src.iter_rows(values_only=True):
        ws_dst.append(row)
        n += 1
        if batch_log and n % batch_log == 0:
            print(f'    ... 已流式写入 {n:,} 行')
    wb_dst.save(out_path)
    return n


def main():
    company_code = '4390'
    output_folder = os.path.join(OUTPUT_ROOT, company_code)
    ensure_dir(output_folder)

    # 查找 4390 最新的 Untested 单文件（不含 _分片_）
    pattern = os.path.join(output_folder, f'{OUTPUT_PREFIX}_Untested_{company_code}_*.xlsx')
    candidates = [
        f for f in glob.glob(pattern)
        if '_分片_' not in os.path.basename(f) and os.path.isfile(f)
    ]
    if not candidates:
        print(f'[ERROR] 未找到 4390 的 Untested 文件: {pattern}')
        sys.exit(1)

    src_path = max(candidates, key=lambda f: (os.path.getmtime(f), f))
    base_name = os.path.basename(src_path).replace('.xlsx', '')

    out_base = os.path.join(output_folder, f'{base_name}_分片')

    # 检查已存在的分片（支持断点续传）
    split_pattern = os.path.join(output_folder, f'{base_name}_分片_*.xlsx')
    existing = glob.glob(split_pattern)
    existing_nums = set()
    for f in existing:
        b = os.path.basename(f)
        try:
            # ..._分片_3.xlsx -> 3
            num = int(b.rsplit('_', 1)[1].replace('.xlsx', ''))
            existing_nums.add(num)
        except (ValueError, IndexError):
            pass

    print(f'读取（流式模式）: {os.path.basename(src_path)}')
    wb_src = load_workbook(src_path, read_only=True, data_only=True)
    sheet_names = wb_src.sheetnames

    # 负开票冲帐 sheet 列表
    neg_sheets = [
        s for s in sheet_names
        if s == '负开票冲帐'
        or (s.startswith('负开票冲帐_') and s[len('负开票冲帐_'):].replace('_', '').isdigit())
    ]

    def _neg_sheet_order(name):
        if name == '负开票冲帐':
            return 0
        suffix = name.split('_', 1)[1] if '_' in name else ''
        return int(suffix) if suffix.isdigit() else 999

    neg_sheets_sorted = sorted(neg_sheets, key=_neg_sheet_order)
    other_sheets = [s for s in ['仅订单', '仅订单及发货单', '仅发货单'] if s in sheet_names]

    written = 0

    # 文件 1：仅订单 + 仅订单及发货单 + 仅发货单 + 负开票冲帐_1
    path_1 = f'{out_base}_1.xlsx'
    if 1 not in existing_nums:
        from openpyxl import Workbook
        wb_1 = Workbook(write_only=True)
        sheets_to_write = list(other_sheets)
        if neg_sheets_sorted:
            sheets_to_write.append(neg_sheets_sorted[0])
        for s in sheets_to_write:
            ws_src = wb_src[s]
            ws_dst = wb_1.create_sheet(s[:31])
            n = 0
            for row in ws_src.iter_rows(values_only=True):
                ws_dst.append(row)
                n += 1
                if n % 500000 == 0 and n:
                    print(f'    ... {s} 已流式写入 {n:,} 行')
            print(f'  已写 {s}: {n:,} 行 -> _分片_1.xlsx')
        wb_1.save(path_1)
        written += 1
    else:
        print('  [跳过] _分片_1.xlsx 已存在')

    # 文件 2~N：负开票冲帐_2, _3, ...
    for i, sheet in enumerate(neg_sheets_sorted[1:], start=2):
        if i in existing_nums:
            print(f'  [跳过] _分片_{i}.xlsx 已存在')
            continue
        path_n = f'{out_base}_{i}.xlsx'
        n = _stream_sheet_to_file(wb_src, sheet, path_n)
        print(f'  已写 {sheet}: {n:,} 行 -> _分片_{i}.xlsx')
        written += 1

    wb_src.close()
    print(f'[OK] 拆分完成，共 {written} 个文件')
    print(f'  主文件（含 仅订单/仅订单及发货单）: {os.path.basename(path_1)}')


if __name__ == '__main__':
    main()
