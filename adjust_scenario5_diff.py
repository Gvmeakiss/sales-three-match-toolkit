# -*- coding: utf-8 -*-
"""
将 汇总_差异分析_全公司.xlsx 中各公司 sheet 的场景5 订单发票金额差异 乘以 0.1，
并对应更新全公司汇总及各家小计/合计。

用法: python3 adjust_scenario5_diff.py
"""

import os
import sys
import re

script_dir = os.path.dirname(os.path.abspath(__file__))
if script_dir not in sys.path:
    sys.path.insert(0, script_dir)

from openpyxl import load_workbook
from config import OUTPUT_ROOT

# 需排除的 sheet（汇总表、分析报告等）
EXCLUDE_SHEETS = {'全公司汇总', '分析报告'}

# 订单发票金额差异列（1-based，含订单发票金额差异时为第7列）
DIFF_COL = 7


def _parse_amt(s):
    """解析 _fmt_amt 格式： 1,234.56 或 (1,234.56) 或 -"""
    if s is None or s == '' or s == '-':
        return 0.0
    text = str(s).strip()
    if not text or text == '-':
        return 0.0
    neg = text.startswith('(') and text.endswith(')')
    if neg:
        text = text[1:-1].strip()
    text = text.replace(',', '')
    try:
        val = float(text)
        return -abs(val) if neg else val
    except ValueError:
        return 0.0


def _fmt_amt(amt):
    """格式化为千分位、2位小数，负数用括号"""
    if amt == 0 or amt != amt:
        return '-'
    if amt < 0:
        return f'({abs(amt):,.2f})'
    return f'{amt:,.2f}'


def _find_scenario_table(ws):
    """定位场景表：返回 (header_row, data_start_row)，表头含 场景编号 且为第一列。"""
    for r in range(1, min(ws.max_row + 1, 50)):
        a1 = ws.cell(row=r, column=1).value
        if a1 is not None and '场景' in str(a1):
            # 检查是否为表头行（第二列含 识别场景 或 数量差异）
            a2 = ws.cell(row=r, column=2).value
            if a2 and '识别' in str(a2):
                return r, r + 1
    return None, None


def _get_scenario_rows(ws, data_start_row):
    """
    从 data_start_row 起解析场景表，返回：
    - scenario5_row: 场景5 所在行号（None 若未找到）
    - subtot_row: 小计 行号
    - total_row: 合计 行号
    """
    scenario5_row = None
    subtot_row = None
    total_row = None
    for r in range(data_start_row, data_start_row + 25):
        col1 = ws.cell(row=r, column=1).value
        if col1 is None:
            continue
        try:
            if int(float(col1)) == 5:
                scenario5_row = r
        except (TypeError, ValueError):
            pass
        if col1 == '小计':
            subtot_row = r
        if col1 == '合计':
            total_row = r
        if scenario5_row is not None and subtot_row is not None and total_row is not None:
            break
    return scenario5_row, subtot_row, total_row


def _has_diff_column(ws, header_row):
    """检查是否有订单发票金额差异列（第7列表头）"""
    h = ws.cell(row=header_row, column=DIFF_COL).value
    return h is not None and '订单' in str(h) and '发票' in str(h)


def process_sheet(ws, sheet_name, is_agg=False):
    """
    处理单个 sheet。
    is_agg: 是否为全公司汇总（汇总表从各公司汇总而来，场景5 = 0.1*旧场景5）
    返回 (modified, old_s5, new_s5) 供全公司汇总用。
    """
    header_row, data_start = _find_scenario_table(ws)
    if header_row is None:
        return False, 0.0, 0.0

    if not _has_diff_column(ws, header_row):
        return False, 0.0, 0.0

    s5_row, subtot_row, total_row = _get_scenario_rows(ws, data_start)
    if s5_row is None or subtot_row is None or total_row is None:
        return False, 0.0, 0.0

    old_s5 = _parse_amt(ws.cell(row=s5_row, column=DIFF_COL).value)
    if is_agg:
        new_s5 = old_s5 * 0.1
    else:
        new_s5 = old_s5 * 0.1

    ws.cell(row=s5_row, column=DIFF_COL, value=_fmt_amt(new_s5))

    # 更新小计、合计：小计_new = 小计_old - 旧场景5 + 新场景5
    old_subtot = _parse_amt(ws.cell(row=subtot_row, column=DIFF_COL).value)
    new_subtot = old_subtot - old_s5 + new_s5
    ws.cell(row=subtot_row, column=DIFF_COL, value=_fmt_amt(new_subtot))
    ws.cell(row=total_row, column=DIFF_COL, value=_fmt_amt(new_subtot))

    return True, old_s5, new_s5


def main():
    output_root = os.path.normpath(os.path.abspath(OUTPUT_ROOT))
    path = os.path.join(output_root, '汇总_差异分析_全公司.xlsx')
    if not os.path.exists(path):
        print(f'[ERROR] 文件不存在: {path}')
        sys.exit(1)

    print(f'加载: {path}')
    wb = load_workbook(path, data_only=False)

    # 1. 处理各公司 sheet（排除全公司汇总、分析报告）
    company_sheets = [s for s in wb.sheetnames if s not in EXCLUDE_SHEETS]
    n_updated = 0
    for name in company_sheets:
        ws = wb[name]
        ok, _, _ = process_sheet(ws, name, is_agg=False)
        if ok:
            n_updated += 1
            print(f'  已更新公司: {name}')

    # 2. 更新全公司汇总
    if '全公司汇总' in wb.sheetnames:
        ok, _, _ = process_sheet(wb['全公司汇总'], '全公司汇总', is_agg=True)
        if ok:
            print('  已更新全公司汇总')
            n_updated += 1

    wb.save(path)
    print(f'[OK] 已保存，共修改 {n_updated} 个 sheet')


if __name__ == '__main__':
    main()
