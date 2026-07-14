# -*- coding: utf-8 -*-
"""
汇总各公司差异分析结果至单一 Excel

在所有公司完成三单匹配及差异分析后运行此脚本，
将每家公司的差异分析按场景明细格式合并至一个 Excel 文件。

输出格式（与 difference_analysis 一致）：
- 全公司汇总 sheet：场景 1-13 + 负开票（场景标号、识别场景、记录数、占比、发票金额、发票金额占比）
- 各公司 sheet：该公司场景明细
- 分析报告 sheet：完全匹配分层统计、各公司明细

使用方法：
    python aggregate_difference_summary.py

输出：
    OutPut/汇总_差异分析_全公司.xlsx
"""

import os
import glob
import sys
from concurrent.futures import ThreadPoolExecutor, as_completed
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font

script_dir = os.path.dirname(os.path.abspath(__file__))
if script_dir not in sys.path:
    sys.path.insert(0, script_dir)

from config import OUTPUT_PREFIX, OUTPUT_ROOT
from utils.path_utils import ensure_dir

def _compute_order_inv_diff_series(df):
    """
    计算订单金额与发票金额的差异。
    订单/发票金额已由 SHKZG 等借贷标识决定正负，不再按发票类型二次取负。
    返回 Series：每行 diff = 订单金额 - 发票金额
    """
    ord_col = '订单金额_本币' if '订单金额_本币' in df.columns else ('订单-金额' if '订单-金额' in df.columns else None)
    inv_col = '发票金额_本币' if '发票金额_本币' in df.columns else ('发票-金额' if '发票-金额' in df.columns else None)
    if ord_col is None or inv_col is None:
        return pd.Series(0.0, index=df.index)
    ord_amt = pd.to_numeric(df[ord_col], errors='coerce').fillna(0)
    inv_amt = pd.to_numeric(df[inv_col], errors='coerce').fillna(0)
    return ord_amt - inv_amt


def _backfill_order_inv_diff_from_detail(company_folder, company_code):
    """
    当场景明细缺少「订单发票金额差异」列时，从差异分析_详细文件读取行级数据计算并回填。
    仅读取场景标号、订单金额、发票金额列以节省内存、提速。
    返回 {sid: total_diff}，若无法计算则返回空 dict。
    """
    pattern = os.path.join(company_folder, f'{OUTPUT_PREFIX}_差异分析_{company_code}_详细*.xlsx')
    detail_files = sorted(glob.glob(pattern))
    if not detail_files:
        return {}
    dfs = []
    for fp in detail_files:
        try:
            header = pd.read_excel(fp, sheet_name='Sheet1', nrows=0)
            cols = set(header.columns)
            if '场景标号' not in cols:
                continue
            usecols = ['场景标号']
            if '订单金额_本币' in cols:
                usecols.append('订单金额_本币')
            elif '订单-金额' in cols:
                usecols.append('订单-金额')
            if '发票金额_本币' in cols:
                usecols.append('发票金额_本币')
            elif '发票-金额' in cols:
                usecols.append('发票-金额')
            if len(usecols) < 3:
                continue
            df = pd.read_excel(fp, sheet_name='Sheet1', usecols=usecols)
            if df.empty:
                continue
            dfs.append(df)
        except Exception:
            continue
    if not dfs:
        return {}
    df = pd.concat(dfs, ignore_index=True) if len(dfs) > 1 else dfs[0]
    diff_series = _compute_order_inv_diff_series(df)
    df = df.copy()
    df['_diff'] = diff_series
    result = {}
    for sid in range(1, 14):
        if sid == 7:
            continue
        mask = df['场景标号'] == sid
        if mask.any():
            result[sid] = float(df.loc[mask, '_diff'].sum())
    return result


def _resolve_output_root():
    """解析输出根目录（Mac 兼容，使用绝对路径）"""
    return os.path.normpath(os.path.abspath(OUTPUT_ROOT))


def _find_company_analysis_file(company_folder):
    """在公司输出目录中查找差异分析统计文件（不含 _详细 的明细文件）。
    多文件同 mtime 时按路径排序取最后一个，保证每次运行选同一文件。"""
    pattern = os.path.join(company_folder, f'{OUTPUT_PREFIX}_差异分析_*.xlsx')
    files = [f for f in glob.glob(pattern) if '_详细' not in os.path.basename(f)]
    if not files:
        return None
    return max(files, key=lambda f: (os.path.getmtime(f), f))


def _match_base(path):
    """
    将分片文件 ..._123_2.xlsx 归组到 ..._123.xlsx。
    主文件 ..._123.xlsx 保持不变。
    """
    b = os.path.basename(path)
    if b.endswith('.xlsx'):
        stem = b[:-5]
        parts = stem.rsplit('_', 1)
        if len(parts) == 2 and parts[1].isdigit():
            # 若去掉最后一段后仍以数字结尾，判定为分片（..._随机后缀_分片号）
            parts2 = parts[0].rsplit('_', 1)
            if len(parts2) == 2 and parts2[1].isdigit():
                return os.path.join(os.path.dirname(path), parts[0] + '.xlsx')
    return path


def _match_part_order(path):
    b = os.path.basename(path)
    stem = b[:-5] if b.endswith('.xlsx') else b
    parts = stem.rsplit('_', 1)
    if len(parts) == 2 and parts[1].isdigit():
        parts2 = parts[0].rsplit('_', 1)
        if len(parts2) == 2 and parts2[1].isdigit():
            return int(parts[1])
    return 0


def _find_company_match_files(company_folder):
    """
    查找公司最新一批主匹配结果文件（含分片）。
    仅匹配 SalesThreeMatchResult_*.xlsx，排除 Untested / 差异分析。
    返回 (base_path, files)。
    """
    pattern = os.path.join(company_folder, f'{OUTPUT_PREFIX}_*.xlsx')
    files = [f for f in glob.glob(pattern) if 'Untested' not in os.path.basename(f) and '差异分析' not in os.path.basename(f)]
    if not files:
        return None, []
    groups = {}
    for f in files:
        base = _match_base(f)
        groups.setdefault(base, []).append(f)
    newest_base = max(groups.keys(), key=lambda b: (os.path.getmtime(b), b))
    group_files = sorted(groups[newest_base], key=_match_part_order)
    return newest_base, group_files


def _find_company_untested_file(company_folder, match_base_path):
    """按主匹配同批次优先定位 Untested；找不到则回退为公司内最新 Untested。
    支持 4390 分片：存在 *_分片_1.xlsx 时优先返回（含 仅订单/仅订单及发货单）。"""
    if not match_base_path:
        fs = glob.glob(os.path.join(company_folder, f'{OUTPUT_PREFIX}_Untested_*.xlsx'))
        split_1 = [f for f in fs if '_分片_1.xlsx' in f]
        if split_1:
            return max(split_1, key=lambda f: (os.path.getmtime(f), f))
        return max(fs, key=lambda f: (os.path.getmtime(f), f)) if fs else None
    base_name = os.path.basename(match_base_path)
    after_prefix = base_name.replace(f'{OUTPUT_PREFIX}_', '', 1).replace('.xlsx', '')
    split_path = os.path.join(company_folder, f'{OUTPUT_PREFIX}_Untested_{after_prefix}_分片_1.xlsx')
    if os.path.exists(split_path):
        return split_path
    same_batch = os.path.join(company_folder, f'{OUTPUT_PREFIX}_Untested_{after_prefix}.xlsx')
    if os.path.exists(same_batch):
        return same_batch
    parts = after_prefix.split('_')
    company_code = parts[0] if parts else ''
    fs = glob.glob(os.path.join(company_folder, f'{OUTPUT_PREFIX}_Untested_{company_code}_*.xlsx'))
    split_1 = [f for f in fs if '_分片_1.xlsx' in f]
    if split_1:
        return max(split_1, key=lambda f: (os.path.getmtime(f), f))
    return max(fs, key=lambda f: (os.path.getmtime(f), f)) if fs else None


def _safe_numeric_series(df, col):
    if col not in df.columns:
        return pd.Series([0.0] * len(df), index=df.index, dtype='float64')
    return pd.to_numeric(df[col], errors='coerce').fillna(0)


# 场景标号与识别场景描述（与 difference_analysis 保持一致）
_SCENARIO_DESC = {
    1: '订单数量=发票≠交货单,订单金额<发票金额',
    2: '订单数量=发票≠交货单,订单金额>发票金额',
    3: '订单数量=发票≠交货单,无差异',
    4: '订单数量≠发票≠交货单,订单金额<发票金额',
    5: '订单数量≠发票≠交货单,订单金额>发票金额',
    6: '订单数量≠发票≠交货单,无差异',
    7: '缺失发票 (无发票无开票金额), N/A',
    8: '无差异,订单金额<发票金额',
    9: '无差异,订单金额>发票金额',
    10: '无差异,无差异',
    11: '订单数量=交货单≠发票,无差异',
    12: '订单数量=交货单≠发票,订单金额<发票金额',
    13: '订单数量=交货单≠发票,订单金额>发票金额',
}


def _scenario_template():
    """1-13 场景定义（与 difference_analysis 一致）"""
    return {
        sid: {'差异笔数': 0, '发票金额': 0.0, '订单发票金额差异': 0.0}
        for sid in [1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13]
    }


def _accumulate_scenarios_from_main(df, scenario_map, tol=0.01):
    if df is None or df.empty:
        return
    ord_qty = _safe_numeric_series(df, '订单数量')
    inv_qty = _safe_numeric_series(df, '发票数量')
    dlv_qty = _safe_numeric_series(df, '交货数量')
    amt_diff = _safe_numeric_series(df, '订单-发票金额差异')
    inv_col = '发票金额_本币' if '发票金额_本币' in df.columns else ('发票-金额' if '发票-金额' in df.columns else None)
    inv_amt = _safe_numeric_series(df, inv_col) if inv_col else pd.Series([0.0] * len(df), index=df.index, dtype='float64')

    ord_inv_eq = (ord_qty - inv_qty).abs() < tol
    dlv_missing = df['交货数量'].isna() if '交货数量' in df.columns else pd.Series([True] * len(df), index=df.index)
    ord_dlv_eq = ~dlv_missing & ((ord_qty - dlv_qty).abs() < tol)
    dlv_inv_eq = (dlv_qty - inv_qty).abs() < tol

    qty_a = ord_inv_eq & (dlv_missing | (~dlv_inv_eq))       # 订单=发票≠交货
    qty_b = ~ord_inv_eq                                       # 订单≠发票≠交货
    qty_c = ord_inv_eq & (~dlv_missing) & dlv_inv_eq         # 三单数量一致
    qty_d = ord_dlv_eq & ~ord_inv_eq                          # 订单=交货≠发票

    amt_lt = amt_diff < -tol
    amt_gt = amt_diff > tol
    amt_eq = ~(amt_lt | amt_gt)

    masks = {
        1: qty_a & amt_lt, 2: qty_a & amt_gt, 3: qty_a & amt_eq,
        4: qty_b & amt_lt, 5: qty_b & amt_gt, 6: qty_b & amt_eq,
        8: qty_c & amt_lt, 9: qty_c & amt_gt, 10: qty_c & amt_eq,
        11: qty_d & amt_eq, 12: qty_d & amt_lt, 13: qty_d & amt_gt,
    }
    for sid, m in masks.items():
        c = int(m.sum())
        if c <= 0:
            continue
        scenario_map[sid]['差异笔数'] += c
        scenario_map[sid]['发票金额'] += float(inv_amt[m].sum())


def _accumulate_missing_invoice_from_untested(untested_path, scenario_map):
    if not untested_path or not os.path.exists(untested_path):
        return
    total_cnt = 0
    total_amt = 0.0
    for sheet in ['仅订单', '仅订单及发货单']:
        try:
            header = pd.read_excel(untested_path, sheet_name=sheet, nrows=0)
            cols = set(header.columns)
            amt_col = '订单金额_本币' if '订单金额_本币' in cols else ('发票金额_本币' if '发票金额_本币' in cols else None)
            usecols = [amt_col] if amt_col else None
            df = pd.read_excel(untested_path, sheet_name=sheet, usecols=usecols)
        except Exception:
            continue
        if df.empty:
            continue
        total_cnt += len(df)
        if amt_col and amt_col in df.columns:
            total_amt += float(pd.to_numeric(df[amt_col], errors='coerce').fillna(0).sum())
    scenario_map[7]['差异笔数'] += int(total_cnt)
    scenario_map[7]['发票金额'] += float(total_amt)


def _compute_main_recalc_from_map(scenario_map):
    """从 scenario_map 计算 main_recalc、missing_invoice（供分析报告用）"""
    s3 = scenario_map.get(3, {})
    s10 = scenario_map.get(10, {})
    main_recalc = {
        '无差异': {
            '记录数': s3.get('差异笔数', 0) + s10.get('差异笔数', 0),
            '发票金额': s3.get('发票金额', 0) + s10.get('发票金额', 0),
        },
        '金额差异': {
            '记录数': scenario_map[1]['差异笔数'] + scenario_map[2]['差异笔数'] + scenario_map[8]['差异笔数'] + scenario_map[9]['差异笔数'] + scenario_map[12]['差异笔数'] + scenario_map[13]['差异笔数'],
            '发票金额': scenario_map[1]['发票金额'] + scenario_map[2]['发票金额'] + scenario_map[8]['发票金额'] + scenario_map[9]['发票金额'] + scenario_map[12]['发票金额'] + scenario_map[13]['发票金额'],
        },
        '数量差异': {
            '记录数': scenario_map[6]['差异笔数'] + scenario_map[11]['差异笔数'],
            '发票金额': scenario_map[6]['发票金额'] + scenario_map[11]['发票金额'],
        },
        '均有差异': {
            '记录数': scenario_map[4]['差异笔数'] + scenario_map[5]['差异笔数'],
            '发票金额': scenario_map[4]['发票金额'] + scenario_map[5]['发票金额'],
        },
    }
    missing_invoice = {'记录数': scenario_map[7]['差异笔数'], '发票金额': scenario_map[7]['发票金额']}
    return {'main_recalc': main_recalc, 'missing_invoice': missing_invoice}


def _compute_company_scenarios(company_folder):
    """计算公司级 13 场景（含场景 7 缺失发票）。"""
    scenario_map = _scenario_template()
    match_base, match_files = _find_company_match_files(company_folder)
    for fp in match_files:
        try:
            header = pd.read_excel(fp, sheet_name='Sheet1', nrows=0)
            cols = set(header.columns)
            usecols = [c for c in ['订单数量', '发票数量', '交货数量', '订单-发票金额差异', '发票金额_本币', '发票-金额'] if c in cols]
            df = pd.read_excel(fp, sheet_name='Sheet1', usecols=usecols if usecols else None)
            _accumulate_scenarios_from_main(df, scenario_map)
        except Exception:
            continue

    untested = _find_company_untested_file(company_folder, match_base)
    _accumulate_missing_invoice_from_untested(untested, scenario_map)

    s3 = scenario_map.get(3, {})
    s10 = scenario_map.get(10, {})
    main_recalc = {
        '无差异': {
            '记录数': s3.get('差异笔数', 0) + s10.get('差异笔数', 0),
            '发票金额': s3.get('发票金额', 0) + s10.get('发票金额', 0),
        },
        '金额差异': {
            '记录数': scenario_map[1]['差异笔数'] + scenario_map[2]['差异笔数'] + scenario_map[8]['差异笔数'] + scenario_map[9]['差异笔数'] + scenario_map[12]['差异笔数'] + scenario_map[13]['差异笔数'],
            '发票金额': scenario_map[1]['发票金额'] + scenario_map[2]['发票金额'] + scenario_map[8]['发票金额'] + scenario_map[9]['发票金额'] + scenario_map[12]['发票金额'] + scenario_map[13]['发票金额'],
        },
        '数量差异': {
            '记录数': scenario_map[6]['差异笔数'] + scenario_map[11]['差异笔数'],
            '发票金额': scenario_map[6]['发票金额'] + scenario_map[11]['发票金额'],
        },
        '均有差异': {
            '记录数': scenario_map[4]['差异笔数'] + scenario_map[5]['差异笔数'],
            '发票金额': scenario_map[4]['发票金额'] + scenario_map[5]['发票金额'],
        },
    }
    missing_invoice = {'记录数': scenario_map[7]['差异笔数'], '发票金额': scenario_map[7]['发票金额']}
    return {'scenarios': scenario_map, 'main_recalc': main_recalc, 'missing_invoice': missing_invoice}


def _read_sheets_fast(path, sheet_names):
    """
    用 openpyxl read_only 模式读取指定 sheet，避免将大表载入内存。
    返回 {sheet_name: DataFrame}，缺失时对应空 DataFrame。
    """
    result = {}
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        for name in sheet_names:
            try:
                ws = wb[name]
                rows = list(ws.iter_rows(values_only=True))
                if not rows:
                    result[name] = pd.DataFrame()
                else:
                    result[name] = pd.DataFrame(rows[1:], columns=rows[0])
            except KeyError:
                result[name] = pd.DataFrame()
    finally:
        wb.close()
    return result


def _safe_val(row, col, default=0):
    """安全获取行中的列值"""
    if col not in row.index:
        return default
    v = pd.to_numeric(row[col], errors='coerce')
    return float(v) if pd.notna(v) else default


def _fmt_cnt(n):
    """记录数格式：千分位，0 时显示 -"""
    if n == 0:
        return '-'
    return f'{int(n):,}'


def _fmt_amt(amt):
    """发票金额格式：千分位、2 位小数，负数用括号"""
    if amt == 0 or amt != amt:
        return '-'
    if amt < 0:
        return f'({abs(amt):,.2f})'
    return f'{amt:,.2f}'


def _write_scenario_table(ws, start_row, scenario_map, neg_inv, hide_order_inv_diff=False):
    """
    按图片格式写入场景表：场景 1-13、小计、14（负开票）、合计。
    列：场景编号、识别场景(数量差异类型,金额差异类型)、记录数、占比、发票金额、发票金额占比；hide_order_inv_diff 时不含订单发票金额差异。
    占比：1-13 及小计相对于小计记录数；14 相对于小计记录数。
    发票金额占比：1-13 及小计相对于小计发票金额；14 相对于合计发票金额。
    小计、合计行加粗；记录数/发票金额为 0 时显示 -；负数金额用括号。
    """
    if hide_order_inv_diff:
        headers = ['场景编号', '识别场景(数量差异类型,金额差异类型)', '记录数', '占比', '发票金额', '发票金额占比']
    else:
        headers = ['场景编号', '识别场景(数量差异类型,金额差异类型)', '记录数', '占比', '发票金额', '发票金额占比', '订单发票金额差异']
    for c, h in enumerate(headers, 1):
        ws.cell(row=start_row, column=c, value=h)
    start_row += 1

    # 先计算小计（1-13 合计）
    subtot_cnt = sum(int(scenario_map.get(s, {}).get('差异笔数', 0)) for s in [1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13])
    subtot_amt = sum(float(scenario_map.get(s, {}).get('发票金额', 0)) for s in [1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13])
    subtot_diff = sum(float(scenario_map.get(s, {}).get('订单发票金额差异', 0)) for s in [1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13])
    neg_cnt = int(neg_inv.get('记录数', 0))
    neg_amt = float(neg_inv.get('发票金额', 0))
    total_amt = subtot_amt + neg_amt

    bold_font = Font(bold=True)

    # 1-13：占比 = cnt/subtot_cnt，发票金额占比 = amt/subtot_amt（场景7 发票金额为 0 用 0.00%）；场景10 不显示
    for sid in [1, 2, 3, 4, 5, 6, 7, 8, 9, 11, 12, 13]:
        s = scenario_map.get(sid, {})
        cnt = int(s.get('差异笔数', 0))
        amt = float(s.get('发票金额', 0))
        order_inv_diff = float(s.get('订单发票金额差异', 0))
        pct_cnt = f'{(cnt / subtot_cnt * 100):.2f}%' if subtot_cnt and subtot_cnt > 0 else '0.00%'
        pct_amt = f'{(amt / subtot_amt * 100):.2f}%' if subtot_amt and abs(subtot_amt) >= 1e-9 and sid != 7 else ('0.00%' if sid == 7 else 'N/A')
        if cnt == 0:
            pct_cnt = '0.00%'
        row = [sid, _SCENARIO_DESC.get(sid, ''), _fmt_cnt(cnt), pct_cnt, _fmt_amt(amt), pct_amt]
        if not hide_order_inv_diff:
            row.append(_fmt_amt(order_inv_diff))
        for c, v in enumerate(row, 1):
            ws.cell(row=start_row, column=c, value=v)
        start_row += 1

    # 小计
    pct_cnt = '100.00%' if subtot_cnt else '0.00%'
    pct_amt = '100.00%' if subtot_amt and abs(subtot_amt) >= 1e-9 else '0.00%'
    row = ['小计', '（1-13 合计）', _fmt_cnt(subtot_cnt), pct_cnt, _fmt_amt(subtot_amt), pct_amt]
    if not hide_order_inv_diff:
        row.append(_fmt_amt(subtot_diff))
    for c, v in enumerate(row, 1):
        cell = ws.cell(row=start_row, column=c, value=v)
        cell.font = bold_font
    start_row += 1

    # 14：占比 = neg_cnt/subtot_cnt（相对于小计）；发票金额占比 = neg_amt/total_amt（相对于合计发票金额）；负开票无订单发票金额差异
    total_cnt = subtot_cnt + neg_cnt
    pct_cnt = f'{(neg_cnt / subtot_cnt * 100):.2f}%' if subtot_cnt and subtot_cnt > 0 else '0.00%'
    pct_amt = f'{(neg_amt / total_amt * 100):.2f}%' if total_amt and abs(total_amt) >= 1e-9 else 'N/A'
    row = [14, '有发票、订单或发运单缺失,Not Test', _fmt_cnt(neg_cnt), pct_cnt, _fmt_amt(neg_amt), pct_amt]
    if not hide_order_inv_diff:
        row.append('-')
    for c, v in enumerate(row, 1):
        ws.cell(row=start_row, column=c, value=v)
    start_row += 1

    # 合计（仅 记录数、发票金额、订单发票金额差异，占比留空，加粗）
    row = ['合计', '（1-13+14）', _fmt_cnt(total_cnt), '', _fmt_amt(total_amt), '']
    if not hide_order_inv_diff:
        row.append(_fmt_amt(subtot_diff))
    for c, v in enumerate(row, 1):
        cell = ws.cell(row=start_row, column=c, value=v)
        cell.font = bold_font
    start_row += 1

    return start_row


def _is_internal_only(df):
    """判断是否无三单匹配结果（仅内部采购）。场景明细格式：场景标号 1-13 的 记录数 之和。"""
    if df is None or df.empty:
        return True
    if '场景标号' not in df.columns:
        return True
    total = 0
    for _, row in df.iterrows():
        sid = row.get('场景标号')
        try:
            n = int(float(sid))
            if 1 <= n <= 13:
                total += _safe_val(row, '记录数')
        except (TypeError, ValueError):
            pass
    return total == 0


def _build_aggregated_scenarios(company_records):
    """
    汇总各公司场景数据，返回 (all_scen, neg_cnt, neg_amt) 供 _write_scenario_table 使用。
    """
    all_scen = _scenario_template()
    neg_cnt = 0
    neg_amt = 0.0
    for r in company_records:
        sm = r.get('scenario_map', {})
        for sid in [1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13]:
            if sid not in sm:
                continue
            all_scen[sid]['差异笔数'] += int(sm[sid].get('差异笔数', 0))
            all_scen[sid]['发票金额'] += float(sm[sid].get('发票金额', 0.0))
            all_scen[sid]['订单发票金额差异'] += float(sm[sid].get('订单发票金额差异', 0.0))
        neg_cnt += int(r.get('neg_inv', {}).get('记录数', 0))
        neg_amt += float(r.get('neg_inv', {}).get('发票金额', 0.0))
    return all_scen, neg_cnt, neg_amt


def _build_analysis_report(company_records):
    """
    构建分析报告：按完全匹配金额占比分组统计，含全量、剔除<80% 两组汇总表（条数+金额及占比），并标注剔除的公司。
    返回 (report_df, narrative_lines, tier_summary, full_table, excl_ge80_table, excl_ge80_companies)
    """
    company_stats = []
    for rec in company_records:
        company_code = rec['company_code']
        mr = rec.get('main_recalc', {})
        match_cnt = float(mr.get('无差异', {}).get('记录数', 0))
        match_amt = float(mr.get('无差异', {}).get('发票金额', 0.0))
        total_cnt = float(sum(v.get('记录数', 0) for v in mr.values()))
        total_amt = float(sum(v.get('发票金额', 0.0) for v in mr.values()))
        neg_inv_amt = float(rec.get('neg_inv', {}).get('发票金额', 0.0))
        if total_amt <= 0 and total_cnt <= 0:
            continue
        pct_amt = (match_amt / total_amt * 100) if total_amt > 0 else 0
        pct_cnt = (match_cnt / total_cnt * 100) if total_cnt > 0 else 0
        if pct_amt >= 80:
            tier = '≥80%'
            remark = ''
        elif pct_amt >= 60:
            tier = '60%-80%'
            remark = ''
        else:
            tier = '<60%'
            remark = '金额匹配率<60%'
        company_stats.append({
            '公司代码': company_code,
            '完全匹配条数': int(match_cnt),
            '总条数': int(total_cnt),
            '完全匹配条数占比': f'{pct_cnt:.1f}%',
            '完全匹配金额': round(match_amt, 2),
            '总发票金额': round(float(total_amt), 2),
            '完全匹配金额占比': f'{pct_amt:.1f}%',
            '分类': tier,
            '备注': remark,
            '负开票金额': round(neg_inv_amt, 2),
        })

    if not company_stats:
        return pd.DataFrame(), [], pd.DataFrame(), pd.DataFrame(), pd.DataFrame(), []

    df = pd.DataFrame(company_stats)
    total_match_amt = df['完全匹配金额'].sum()
    total_all_amt = df['总发票金额'].sum()
    total_match_cnt = df['完全匹配条数'].sum()
    total_all_cnt = df['总条数'].sum()
    overall_pct_amt = (total_match_amt / total_all_amt * 100) if total_all_amt > 0 else 0
    overall_pct_cnt = (total_match_cnt / total_all_cnt * 100) if total_all_cnt > 0 else 0

    pct_vals = df['完全匹配金额占比'].str.replace('%', '', regex=False).astype(float)
    df_ge80 = df[pct_vals >= 80]
    excl_ge80_companies = df[pct_vals < 80]['公司代码'].tolist()

    # 分层汇总（含条数、金额及占比）
    def _tier_row(tier_df, label):
        if tier_df.empty:
            return {'分类': label, '公司数': 0, '完全匹配条数': 0, '总条数': 0, '条数占比': 'N/A',
                    '完全匹配金额': 0.0, '总发票金额': 0.0, '金额占比': 'N/A'}
        m_cnt = tier_df['完全匹配条数'].sum()
        t_cnt = tier_df['总条数'].sum()
        m_amt = tier_df['完全匹配金额'].sum()
        t_amt = tier_df['总发票金额'].sum()
        pct_c = f'{(m_cnt / t_cnt * 100):.1f}%' if t_cnt > 0 else 'N/A'
        pct_a = f'{(m_amt / t_amt * 100):.1f}%' if t_amt > 0 else 'N/A'
        return {'分类': label, '公司数': len(tier_df), '完全匹配条数': int(m_cnt), '总条数': int(t_cnt), '条数占比': pct_c,
                '完全匹配金额': round(m_amt, 2), '总发票金额': round(t_amt, 2), '金额占比': pct_a}

    tier_summary = pd.DataFrame([
        _tier_row(df[df['分类'] == '≥80%'], '≥80%'),
        _tier_row(df[df['分类'] == '60%-80%'], '60%-80%'),
        _tier_row(df[df['分类'] == '<60%'], '<60%'),
    ])

    # 全量汇总表
    full_table = pd.DataFrame([{
        '口径': '全量（所有公司）',
        '保留公司数': len(df),
        '完全匹配条数': int(total_match_cnt),
        '总条数': int(total_all_cnt),
        '条数占比': f'{overall_pct_cnt:.1f}%',
        '完全匹配金额': round(total_match_amt, 2),
        '总发票金额': round(total_all_amt, 2),
        '金额占比': f'{overall_pct_amt:.1f}%',
    }])

    # 剔除<80%后的汇总表（仅保留≥80%公司）
    excl_ge80_match_cnt = df_ge80['完全匹配条数'].sum()
    excl_ge80_total_cnt = df_ge80['总条数'].sum()
    excl_ge80_match_amt = df_ge80['完全匹配金额'].sum()
    excl_ge80_total_amt = df_ge80['总发票金额'].sum()
    excl_ge80_pct_cnt = (excl_ge80_match_cnt / excl_ge80_total_cnt * 100) if excl_ge80_total_cnt > 0 else 0
    excl_ge80_pct_amt = (excl_ge80_match_amt / excl_ge80_total_amt * 100) if excl_ge80_total_amt > 0 else 0
    excl_ge80_table = pd.DataFrame([{
        '口径': '剔除匹配率<80%的公司后',
        '保留公司数': len(df_ge80),
        '完全匹配条数': int(excl_ge80_match_cnt),
        '总条数': int(excl_ge80_total_cnt),
        '条数占比': f'{excl_ge80_pct_cnt:.1f}%',
        '完全匹配金额': round(excl_ge80_match_amt, 2),
        '总发票金额': round(excl_ge80_total_amt, 2),
        '金额占比': f'{excl_ge80_pct_amt:.1f}%',
    }])

    narrative = [
        '',
        '【数据说明】',
        '',
        '一、口径定义',
        f'本报告汇总 {len(company_stats)} 家公司的销售三单匹配结果。按场景明细（1-13 + 负开票）统计；完全匹配（无差异）：订单与发票数量一致且金额无差异（含交货差异场景3与完全一致场景10）。',
        '',
        '二、全量数据',
        f'完全匹配条数 {total_match_cnt:,.0f} 条（条数占比 {overall_pct_cnt:.1f}%），完全匹配金额 {total_match_amt:,.2f} 元（金额占比 {overall_pct_amt:.1f}%）。',
        '',
        '三、敏感性分析（剔除匹配率<80%的公司后）',
        f'剔除匹配率<80%的公司后：保留 {len(df_ge80)} 家，条数占比 {excl_ge80_pct_cnt:.1f}%，金额占比 {excl_ge80_pct_amt:.1f}%。',
        f'剔除公司（金额匹配率<80%）共 {len(excl_ge80_companies)} 家：{", ".join(str(c) for c in excl_ge80_companies) if excl_ge80_companies else "无"}。',
        '',
        '四、条数与金额匹配率差异说明',
        '三单匹配按订单行关联，实现订单、交货、发票逐行对应。条数匹配率反映可成功建立对应关系的条目占比，金额匹配率反映对应条目中金额完全一致的比例。当条数匹配率较高而金额匹配率较低时，表明匹配方法可覆盖较多业务条目，部分已匹配条目的订单金额与发票金额存在差异，此为数据中的客观事实。',
    ]
    return df, narrative, tier_summary, full_table, excl_ge80_table, excl_ge80_companies


def _process_single_company(company_code, analysis_path, skip_order_inv_diff=False):
    """
    处理单家公司的差异分析，返回 (record, backfill_msg)。
    record 为 None 表示失败；backfill_msg 非空时表示发生了回填。
    skip_order_inv_diff=True 时跳过订单发票金额差异计算与回填，该列填 0。
    """
    try:
        sheets = _read_sheets_fast(analysis_path, ['场景明细'])
        df_scenario = sheets.get('场景明细')
        if df_scenario is None:
            df_scenario = pd.DataFrame()
        if df_scenario.empty:
            return None, f'[WARNING] 公司 {company_code} 汇总表为空，跳过'
        scenario_map = _scenario_template()
        neg_cnt = 0
        neg_amt = 0.0
        for _, row in df_scenario.iterrows():
            sid = row.get('场景标号')
            if sid is None or (isinstance(sid, float) and pd.isna(sid)):
                continue
            if isinstance(sid, str) and sid.strip() == '负开票':
                neg_cnt = int(_safe_val(row, '记录数'))
                neg_amt = float(_safe_val(row, '发票金额'))
                continue
            try:
                n = int(float(sid))
                if 1 <= n <= 13:
                    scenario_map[n]['差异笔数'] += int(_safe_val(row, '记录数'))
                    scenario_map[n]['发票金额'] += float(_safe_val(row, '发票金额'))
                    if not skip_order_inv_diff and '订单发票金额差异' in df_scenario.columns:
                        scenario_map[n]['订单发票金额差异'] += float(_safe_val(row, '订单发票金额差异'))
            except (TypeError, ValueError):
                pass
        backfill_msg = None
        if not skip_order_inv_diff and '订单发票金额差异' not in df_scenario.columns:
            company_folder = os.path.dirname(analysis_path)
            backfill = _backfill_order_inv_diff_from_detail(company_folder, company_code)
            if backfill:
                for sid, val in backfill.items():
                    scenario_map[sid]['订单发票金额差异'] = round(val, 2)
                backfill_msg = f'  [回填] 公司 {company_code}: 从详细文件补充订单发票金额差异'
        scen = _compute_main_recalc_from_map(scenario_map)
        record = {
            'company_code': company_code,
            'df_scenario': df_scenario,
            'scenario_map': scenario_map,
            'main_recalc': scen['main_recalc'],
            'missing_invoice': scen['missing_invoice'],
            'neg_inv': {'记录数': neg_cnt, '发票金额': neg_amt},
        }
        return record, backfill_msg
    except Exception as e:
        return None, f'[WARNING] 公司 {company_code} 读取失败: {e}'


def main():
    import argparse
    parser = argparse.ArgumentParser(description='汇总各公司差异分析结果')
    parser.add_argument('--limit', type=int, default=None, help='仅处理前 N 家公司（用于快速校验格式）')
    parser.add_argument('--companies', type=str, default=None, help='仅处理指定公司，逗号分隔，如 4010,4030')
    parser.add_argument('--workers', type=int, default=None, help='并行线程数（默认 cpu 核心数，0=单线程）')
    parser.add_argument('--no-order-inv-diff', action='store_true', help='不计算订单发票金额差异，跳过回填详细文件，该列填 0，输出更快')
    args = parser.parse_args()

    print('汇总各公司差异分析结果...')
    output_root = _resolve_output_root()
    if not os.path.exists(output_root):
        print(f'[ERROR] 输出根目录不存在: {output_root}')
        return

    companies = []
    for name in sorted(os.listdir(output_root)):
        path = os.path.join(output_root, name)
        if os.path.isdir(path) and name.strip().isdigit():
            analysis_file = _find_company_analysis_file(path)
            if analysis_file:
                companies.append((name, analysis_file))
            else:
                print(f'[SKIP] 公司 {name}: 未找到差异分析文件')

    if not companies:
        print('[WARNING] 未找到任何公司的差异分析文件')
        return

    if args.companies:
        want = {s.strip() for s in args.companies.split(',') if s.strip()}
        companies = [(c, p) for c, p in companies if c in want]
        print(f'[INFO] --companies，仅处理: {sorted(want)}')
    elif args.limit:
        companies = companies[: args.limit]
        print(f'[INFO] --limit {args.limit}，仅处理前 {len(companies)} 家')

    if args.limit or args.companies:
        output_file = os.path.join(output_root, '汇总_差异分析_格式预览.xlsx')
    elif args.no_order_inv_diff:
        output_file = os.path.join(output_root, '汇总_差异分析_全公司_无订单发票差异.xlsx')
    else:
        output_file = os.path.join(output_root, '汇总_差异分析_全公司.xlsx')
    ensure_dir(output_root)
    n_workers = args.workers
    if n_workers is None:
        n_workers = min(os.cpu_count() or 4, len(companies), 16)
    elif n_workers <= 0:
        n_workers = 1
    skip_diff = getattr(args, 'no_order_inv_diff', False)
    print(f'共 {len(companies)} 家公司，输出: {output_file}' + (f'，{n_workers} 线程并行' if n_workers > 1 else '') + ('，跳过订单发票金额差异' if skip_diff else ''))

    # 直接使用差异分析文件的场景明细（含发运单数量修正后的结果），不再从原始匹配文件重算
    company_records = []
    backfill_msgs = []
    if n_workers <= 1:
        for company_code, analysis_path in companies:
            record, msg = _process_single_company(company_code, analysis_path, skip_order_inv_diff=skip_diff)
            if record:
                company_records.append(record)
            if record is None and msg:
                print(msg)
            if msg and '回填' in (msg or ''):
                backfill_msgs.append(msg)
    else:
        with ThreadPoolExecutor(max_workers=n_workers) as executor:
            futures = {executor.submit(_process_single_company, c, p, skip_diff): c for c, p in companies}
            # 按 companies 顺序收集结果
            results_by_code = {}
            for future in as_completed(futures):
                company_code = futures[future]
                try:
                    record, msg = future.result()
                    results_by_code[company_code] = (record, msg)
                except Exception as e:
                    results_by_code[company_code] = (None, f'[WARNING] 公司 {company_code} 读取失败: {e}')
            for company_code, _ in companies:
                record, msg = results_by_code.get(company_code, (None, None))
                if record:
                    company_records.append(record)
                if msg:
                    if record is None:
                        print(msg)
                    elif '回填' in msg:
                        backfill_msgs.append(msg)
    for m in backfill_msgs:
        print(m)

    if not company_records:
        print('[WARNING] 无有效数据可汇总')
        return

    # 构建全公司汇总（按图片格式：场景 1-13、小计、14、合计）
    agg_scen, agg_neg_cnt, agg_neg_amt = _build_aggregated_scenarios(company_records)
    # 构建审计分析报告
    report_df, narrative_lines, tier_summary, full_table, excl_ge80_table, excl_ge80_companies = _build_analysis_report(company_records)

    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        # 1. 全公司汇总 sheet（置于最前，按图片格式：场景编号、识别场景、小计、合计）
        ws_summary = writer.book.create_sheet('全公司汇总', 0)
        writer.sheets['全公司汇总'] = ws_summary
        row = 1
        ws_summary.cell(row=row, column=1, value='【场景明细】')
        row += 1
        _write_scenario_table(ws_summary, row, agg_scen, {'记录数': agg_neg_cnt, '发票金额': agg_neg_amt}, hide_order_inv_diff=skip_diff)

        # 2. 分析报告 sheet（三组汇总表、分层汇总、各公司明细、数据说明）
        ws_report = writer.book.create_sheet('分析报告', 1)
        writer.sheets['分析报告'] = ws_report
        r = 1
        # 两组汇总表：全量、剔除<80%（含条数、金额及占比），并标注剔除公司
        ws_report.cell(row=r, column=1, value='【汇总表一】全量（所有公司）')
        r += 1
        if not full_table.empty:
            full_table.to_excel(writer, sheet_name='分析报告', index=False, startrow=r, startcol=0)
            r += len(full_table) + 2
        ws_report.cell(row=r, column=1, value='【汇总表二】剔除匹配率<80%的公司后（仅保留≥80%公司）')
        r += 1
        if not excl_ge80_table.empty:
            excl_ge80_table.to_excel(writer, sheet_name='分析报告', index=False, startrow=r, startcol=0)
            r += len(excl_ge80_table) + 2
        if excl_ge80_companies:
            ws_report.cell(row=r, column=1, value=f'剔除公司（金额匹配率<80%）：{", ".join(str(c) for c in excl_ge80_companies)}')
            r += 1
        ws_report.cell(row=r, column=1, value='【分层汇总】完全匹配金额占比分组')
        r += 1
        if not tier_summary.empty:
            tier_summary.to_excel(writer, sheet_name='分析报告', index=False, startrow=r, startcol=0)
            r += len(tier_summary) + 2
        ws_report.cell(row=r, column=1, value='【各公司明细】')
        r += 1
        if not report_df.empty:
            report_df.to_excel(writer, sheet_name='分析报告', index=False, startrow=r, startcol=0)
            r += len(report_df) + 2
        for line in narrative_lines:
            ws_report.cell(row=r, column=1, value=line)
            r += 1

        # 3. 各公司 sheet（按图片格式：场景 1-13、小计、14、合计）
        for rec in company_records:
            company_code = rec['company_code']
            df_scenario = rec.get('df_scenario')
            if df_scenario is None:
                df_scenario = pd.DataFrame()
            scenario_map = rec['scenario_map']
            neg_inv = rec['neg_inv']
            sheet_name = str(company_code)
            if _is_internal_only(df_scenario):
                sheet_name = f'{company_code}（仅内部采购）'
            if len(sheet_name) > 31:
                sheet_name = sheet_name[:31]
            try:
                ws = writer.book.create_sheet(sheet_name)
                writer.sheets[sheet_name] = ws
                next_row = 1
                ws.cell(row=next_row, column=1, value='【场景明细】')
                next_row += 1
                _write_scenario_table(ws, next_row, scenario_map, neg_inv, hide_order_inv_diff=skip_diff)
            except Exception as e:
                print(f'[WARNING] 公司 {company_code} 写入失败: {e}')

    print(f'[OK] 汇总完成: {output_file}')


if __name__ == '__main__':
    main()
